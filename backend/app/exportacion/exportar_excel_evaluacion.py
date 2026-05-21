"""
Exportación a Excel de la Matriz de Evaluación (CACTUCITOS-Evaluacion.xlsx).
Replica exactamente la estructura de columnas de la plantilla original.
"""

from __future__ import annotations
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..config import SALIDAS_DIR, EVALUACION_COLUMNS
from ..models import Hallazgo, ContextoEntidad


def generar_excel_evaluacion(
    hallazgos: list[Hallazgo],
    contexto: ContextoEntidad,
) -> Path | None:
    """
    Genera el Excel de evaluación con la estructura exacta de la plantilla.

    Columnas: No, HALLAZGO, CRITERIO, CAUSA/EFECTO, CONCLUSIÓN, P, I, RIESGO,
    NIVEL RIESGO, RECOMENDACIÓN.

    Args:
        hallazgos: Lista de hallazgos validados.
        contexto: Contexto de la entidad.

    Returns:
        Ruta al archivo generado.
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Evaluación"

        # ── Estilos ──────────────────────────────────────────────
        font_titulo = Font(name="Calibri", size=12, bold=True, color="003366")
        font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        font_normal = Font(name="Calibri", size=9)
        font_label = Font(name="Calibri", size=10, bold=True)

        fill_header = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        fill_alt = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        border_thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        align_wrap = Alignment(wrap_text=True, vertical="top")
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # ── Encabezado del documento (filas 1-6) ─────────────────
        grupo = contexto.grupo_auditor or "Grupo Auditor"
        integrantes_str = ", ".join(contexto.integrantes) if contexto.integrantes else ""
        alcance = contexto.alcance or "Evaluar el control interno de la entidad"

        # Fila 1: Grupo y Banco
        ws.merge_cells("B1:G1")
        ws["B1"] = f"GRUPO: {grupo}"
        ws["B1"].font = font_label
        ws.merge_cells("H1:N1")
        ws["H1"] = f"BANCO AUDITADO: {contexto.nombre_entidad}"
        ws["H1"].font = font_label

        # Fila 2: Integrantes
        ws.merge_cells("B2:N2")
        ws["B2"] = f"Nombres y Apellidos: {integrantes_str}"
        ws["B2"].font = font_normal

        # Fila 3: Alcance
        ws.merge_cells("B3:N3")
        ws["B3"] = f"Alcance: {alcance}"
        ws["B3"].font = font_normal
        ws["B3"].alignment = align_wrap

        # Fila 4: Periodo
        ws.merge_cells("B4:N4")
        ws["B4"] = f"Periodo evaluado: {contexto.periodo_evaluado}"
        ws["B4"].font = font_normal

        # Fila 5: Título
        ws.merge_cells("A5:N5")
        ws["A5"] = "MATRIZ DE EVALUACIÓN DE RIESGOS"
        ws["A5"].font = font_titulo
        ws["A5"].alignment = Alignment(horizontal="center")

        # Fila 6: Vacía (separador)

        # ── Fila 7: Headers de columnas ──────────────────────────
        headers = {
            1: ("No.", 5),
            2: ("HALLAZGO", 50),
            7: ("CRITERIO", 35),
            8: ("CAUSA / EFECTO", 35),
            9: ("CONCLUSIÓN", 35),
            10: ("P", 5),
            11: ("I", 5),
            12: ("RIESGO", 8),
            13: ("NIVEL RIESGO", 12),
            14: ("RECOMENDACIÓN", 40),
        }

        for col, (titulo, ancho) in headers.items():
            cell = ws.cell(row=7, column=col, value=titulo)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_thin
            ws.column_dimensions[get_column_letter(col)].width = ancho

        # Merge columnas B-F para HALLAZGO (header)
        ws.merge_cells("B7:F7")

        # Columnas C-F para relleno del header
        for col in range(3, 7):
            cell = ws.cell(row=7, column=col)
            cell.fill = fill_header
            cell.border = border_thin

        # ── Filas de datos (a partir de fila 8) ──────────────────
        for idx, hallazgo in enumerate(hallazgos):
            fila = 8 + idx

            # Construir texto de hallazgo con citas
            texto_hallazgo = f"R{idx + 1}. {hallazgo.descripcion}"
            if hallazgo.citas_bibliograficas:
                citas_texto = "\n".join(
                    f"- {c.documento}, {c.seccion}"
                    for c in hallazgo.citas_bibliograficas
                )
                texto_hallazgo += f"\n{citas_texto}"

            # Causa / Efecto combinado
            causa_efecto = (
                f"Causa: {hallazgo.causa}\n"
                f"Efecto: {hallazgo.efecto}"
            )

            # Escribir datos
            datos_fila = {
                1: idx + 1,
                2: texto_hallazgo,
                7: hallazgo.criterio,
                8: causa_efecto,
                9: hallazgo.conclusion,
                10: hallazgo.probabilidad,
                11: hallazgo.impacto,
                14: hallazgo.recomendacion,
            }

            for col, valor in datos_fila.items():
                cell = ws.cell(row=fila, column=col, value=valor)
                cell.font = font_normal
                cell.alignment = align_wrap
                cell.border = border_thin

                # Alternar color de fondo
                if idx % 2 == 1:
                    cell.fill = fill_alt

            # Merge B-F para la columna HALLAZGO
            ws.merge_cells(
                start_row=fila, start_column=2,
                end_row=fila, end_column=6
            )
            for col in range(3, 7):
                cell = ws.cell(row=fila, column=col)
                cell.border = border_thin
                if idx % 2 == 1:
                    cell.fill = fill_alt

            # Fórmulas para RIESGO y NIVEL RIESGO
            col_p = get_column_letter(10)
            col_i = get_column_letter(11)

            # RIESGO = P * I
            ws.cell(row=fila, column=12).value = f"=+{col_p}{fila}*{col_i}{fila}"
            ws.cell(row=fila, column=12).font = font_normal
            ws.cell(row=fila, column=12).alignment = align_center
            ws.cell(row=fila, column=12).border = border_thin

            # NIVEL RIESGO (fórmula condicional exacta de la plantilla)
            col_r = get_column_letter(12)
            formula_nivel = (
                f'=IF({col_r}{fila}<=2,"Muy Bajo",'
                f'IF(AND({col_r}{fila}>=3,{col_r}{fila}<=4),"Bajo",'
                f'IF(AND({col_r}{fila}>=5,{col_r}{fila}<=9),"Medio",'
                f'IF(AND({col_r}{fila}>9,{col_r}{fila}<20),"Alto",'
                f'IF({col_r}{fila}>=20,"Extremo","Valores errados")))))'
            )
            ws.cell(row=fila, column=13).value = formula_nivel
            ws.cell(row=fila, column=13).font = font_normal
            ws.cell(row=fila, column=13).alignment = align_center
            ws.cell(row=fila, column=13).border = border_thin

            # Centrar P e I
            ws.cell(row=fila, column=10).alignment = align_center
            ws.cell(row=fila, column=11).alignment = align_center

        # ── Ajustar altura de filas ──────────────────────────────
        for fila in range(8, 8 + len(hallazgos)):
            ws.row_dimensions[fila].height = 80

        # ── Guardar ──────────────────────────────────────────────
        nombre_archivo = f"Evaluacion_{contexto.nombre_entidad.replace(' ', '_')}.xlsx"
        ruta_salida = SALIDAS_DIR / nombre_archivo
        wb.save(str(ruta_salida))
        print(f"  ✅ Excel de evaluación generado: {ruta_salida.name}")
        return ruta_salida

    except Exception as e:
        print(f"  ❌ Error generando Excel de evaluación: {e}")
        return None
