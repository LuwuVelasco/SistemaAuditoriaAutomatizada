"""
Exportación a Excel del Informe COSO de Control Interno.
Replica la estructura de la plantilla '13. Informe COSO control interno...xlsx'.
"""

from __future__ import annotations
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..config import SALIDAS_DIR
from ..models import ComponenteCOSO, ContextoEntidad


def generar_excel_coso(
    componentes: list[ComponenteCOSO],
    contexto: ContextoEntidad,
) -> Path | None:
    """
    Genera el Excel del Informe COSO de Control Interno.

    Estructura: Componente → Lineamientos → Cantidad de deficiencias → Observaciones

    Args:
        componentes: Lista de componentes COSO evaluados.
        contexto: Contexto de la entidad.

    Returns:
        Ruta al archivo generado.
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "COSO-CI"

        # ── Estilos ──────────────────────────────────────────────
        font_titulo_grande = Font(name="Calibri", size=14, bold=True, color="003366")
        font_titulo = Font(name="Calibri", size=12, bold=True, color="003366")
        font_subtitulo = Font(name="Calibri", size=11, bold=True)
        font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        font_normal = Font(name="Calibri", size=9)
        font_componente = Font(name="Calibri", size=10, bold=True, color="003366")
        font_label = Font(name="Calibri", size=10, bold=True)

        fill_header = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        fill_componente = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        fill_deficiencia = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

        border_thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        align_wrap = Alignment(wrap_text=True, vertical="top")
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # ── Encabezado (filas 1-7) ───────────────────────────────
        grupo = contexto.grupo_auditor or "Grupo Auditor"

        # Fila 2: Grupo y número de informe
        ws.merge_cells("E2:G2")
        ws["E2"] = f"NOMBRE GRUPO: {grupo}"
        ws["E2"].font = font_label
        ws["H2"] = f"No. Informe\nSIS-324-001/2026"
        ws["H2"].font = font_normal
        ws["H2"].alignment = align_wrap

        # Fila 3: Título principal
        ws.merge_cells("E3:H3")
        ws["E3"] = "INFORME EVALUACIÓN INDEPENDIENTE"
        ws["E3"].font = font_titulo_grande

        # Fila 4: Subtítulo
        ws.merge_cells("E4:H4")
        ws["E4"] = "INFORME DE SEGUIMIENTO AL SISTEMA DE CONTROL INTERNO"
        ws["E4"].font = font_titulo

        # Fila 5: Periodo
        ws.merge_cells("B5:H5")
        ws["B5"] = f"PERIODO EVALUADO: {contexto.periodo_evaluado}"
        ws["B5"].font = font_label

        # Fila 6: Descripción del alcance
        ws.merge_cells("B6:H6")
        ws["B6"] = (
            f"Sistema de Control Interno de {contexto.nombre_entidad}, "
            f"evaluación de los componentes: Ambiente de Control, "
            f"Evaluación de Riesgos, Actividades de Control, "
            f"Información y Comunicación, Actividades de Monitoreo"
        )
        ws["B6"].font = font_normal
        ws["B6"].alignment = align_wrap

        # Fila 7: Subtítulo tabla
        ws.merge_cells("B7:H7")
        ws["B7"] = "SEGUIMIENTO A LOS COMPONENTES Y LINEAMIENTOS"
        ws["B7"].font = font_subtitulo

        # ── Fila 8: Headers de la tabla ──────────────────────────
        headers_tabla = {
            "B": ("No.", 5),
            "C": ("COMPONENTE", 30),
            "F": ("LINEAMIENTO", 50),
            "G": ("Cantidad de deficiencias encontradas", 15),
            "H": ("Descripción de observaciones y deficiencias", 60),
        }

        for col_letter, (titulo, ancho) in headers_tabla.items():
            cell = ws[f"{col_letter}8"]
            cell.value = titulo
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_thin
            ws.column_dimensions[col_letter].width = ancho

        # Merge C-E para COMPONENTE header
        ws.merge_cells("C8:E8")
        for col in ["D", "E"]:
            cell = ws[f"{col}8"]
            cell.fill = fill_header
            cell.border = border_thin

        # ── Filas de datos ───────────────────────────────────────
        fila_actual = 9

        for componente in componentes:
            fila_inicio_componente = fila_actual
            primera_fila_componente = True

            for lin in componente.lineamientos:
                # Col B: Número del componente (solo en primera fila)
                if primera_fila_componente:
                    ws.cell(row=fila_actual, column=2, value=float(componente.numero))
                    ws.cell(row=fila_actual, column=2).font = font_componente
                    ws.cell(row=fila_actual, column=2).alignment = align_center
                    ws.cell(row=fila_actual, column=2).fill = fill_componente
                    ws.cell(row=fila_actual, column=2).border = border_thin

                    # Col C-E: Nombre del componente
                    ws.cell(row=fila_actual, column=3, value=componente.nombre)
                    ws.cell(row=fila_actual, column=3).font = font_componente
                    ws.cell(row=fila_actual, column=3).fill = fill_componente
                    ws.cell(row=fila_actual, column=3).alignment = align_wrap
                    ws.cell(row=fila_actual, column=3).border = border_thin

                    primera_fila_componente = False
                else:
                    # Celdas vacías para merge visual
                    for col in [2, 3]:
                        cell = ws.cell(row=fila_actual, column=col)
                        cell.border = border_thin

                # Col D-E: borders para merge visual
                for col in [4, 5]:
                    cell = ws.cell(row=fila_actual, column=col)
                    cell.border = border_thin

                # Col F: Lineamiento
                ws.cell(row=fila_actual, column=6, value=lin.descripcion)
                ws.cell(row=fila_actual, column=6).font = font_normal
                ws.cell(row=fila_actual, column=6).alignment = align_wrap
                ws.cell(row=fila_actual, column=6).border = border_thin

                # Col G: Cantidad de deficiencias
                if lin.cantidad_deficiencias > 0:
                    ws.cell(row=fila_actual, column=7, value=float(lin.cantidad_deficiencias))
                    ws.cell(row=fila_actual, column=7).fill = fill_deficiencia
                else:
                    ws.cell(row=fila_actual, column=7, value="")

                ws.cell(row=fila_actual, column=7).font = font_normal
                ws.cell(row=fila_actual, column=7).alignment = align_center
                ws.cell(row=fila_actual, column=7).border = border_thin

                # Col H: Observaciones (solo en primera fila del componente)
                # La plantilla original pone todas las observaciones en la primera fila
                if fila_actual == fila_inicio_componente:
                    # Recopilar todas las observaciones del componente
                    todas_obs = "\n\n".join(
                        l.observaciones for l in componente.lineamientos
                        if l.observaciones
                    )
                    ws.cell(row=fila_actual, column=8, value=todas_obs)
                    ws.cell(row=fila_actual, column=8).font = font_normal
                    ws.cell(row=fila_actual, column=8).alignment = align_wrap
                    ws.cell(row=fila_actual, column=8).border = border_thin
                else:
                    ws.cell(row=fila_actual, column=8).border = border_thin

                # Altura de fila
                ws.row_dimensions[fila_actual].height = 60

                fila_actual += 1

            # Merge celdas del componente si tiene múltiples lineamientos
            if len(componente.lineamientos) > 1:
                ws.merge_cells(
                    start_row=fila_inicio_componente, start_column=2,
                    end_row=fila_actual - 1, end_column=2
                )
                ws.merge_cells(
                    start_row=fila_inicio_componente, start_column=3,
                    end_row=fila_actual - 1, end_column=5
                )
                # Merge observaciones
                ws.merge_cells(
                    start_row=fila_inicio_componente, start_column=8,
                    end_row=fila_actual - 1, end_column=8
                )

        # ── Conclusiones ─────────────────────────────────────────
        fila_actual += 1
        ws.merge_cells(
            start_row=fila_actual, start_column=2,
            end_row=fila_actual, end_column=8
        )
        ws.cell(row=fila_actual, column=2, value="Conclusiones del Estado del Sistema de Control Interno")
        ws.cell(row=fila_actual, column=2).font = font_subtitulo

        # Firmantes
        fila_actual += 1
        if contexto.integrantes:
            nombres = ", ".join(contexto.integrantes)
            ws.cell(row=fila_actual, column=2, value=f"- {nombres}")
            ws.cell(row=fila_actual, column=2).font = font_normal

        # ── Guardar ──────────────────────────────────────────────
        nombre_archivo = (
            f"Informe_COSO_Control_Interno_"
            f"{contexto.nombre_entidad.replace(' ', '_')}.xlsx"
        )
        ruta_salida = SALIDAS_DIR / nombre_archivo
        wb.save(str(ruta_salida))
        print(f"  ✅ Informe COSO generado: {ruta_salida.name}")
        return ruta_salida

    except Exception as e:
        print(f"  ❌ Error generando Excel COSO: {e}")
        return None
